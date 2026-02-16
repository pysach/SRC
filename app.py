from flask import send_from_directory
from flask import Flask, render_template, request, redirect, url_for
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import WebDriverException

import os
import time
import requests
import openpyxl



app = Flask(__name__)

# Route to serve files from uploads folder
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def open_urls_in_tabs(file_path, column_name):
    chrome_options = Options()
    chrome_options.add_argument("--window-size=1920,1080")
    # chromedriver_path = "C:/DRIVER/chromedriver/chromedriver.exe"
    # service = Service(chromedriver_path)
    # driver = webdriver.Chrome(service=service, options=chrome_options)
    driver = webdriver.Firefox()
    screenshots_folder = os.path.join(UPLOAD_FOLDER, 'screenshots')
    os.makedirs(screenshots_folder, exist_ok=True)
    screenshot_data = []
    wb_cookies = openpyxl.load_workbook("./keys/cookies.xlsx")

    try:
        df = pd.read_excel(file_path)
        # if 'links' not in df.columns:
        #     print(f"Column '{column_name}' not found.")
        #     driver.quit()
        #     return []

        for index, row in df.iterrows():
            print(row)
            if isinstance(row['url'], str) and row['url'].strip():
                # Get HTTP response code
                try:
                    response = requests.get(row['url'], timeout=5)
                    status_code = response.status_code
                except Exception as e:
                    status_code = 'ERR'
                if index > 0:
                    driver.execute_script("window.open('');")
                    driver.switch_to.window(driver.window_handles[-1])
                driver.get(row['url'])
                if row['OTT'] in ['X','FACEBOOK','INSTAGRAM']:
                    ws_cookies = wb_cookies[row['OTT']]
                    assert isinstance(ws_cookies.max_row, object)
                    i_cook_limit = int(ws_cookies.max_row)
                    for i_cook in range(2, i_cook_limit+1):
                        cook_name = ws_cookies['A' + str(i_cook)].value
                        cook_val = ws_cookies['B' + str(i_cook)].value
                        cook_domain = ws_cookies['C' + str(i_cook)].value
                        cook_path = ws_cookies['D' + str(i_cook)].value
                        driver.add_cookie({"name": str(cook_name), "value": str(cook_val), "domain": str(cook_domain),"Path": str(cook_path)})
                    driver.refresh()


                time.sleep(2)
                screenshot_path = os.path.join(screenshots_folder, f'screenshot_{index+1}.png')
                driver.save_screenshot(screenshot_path)
                screenshot_data.append({
                    'url': row['url'],
                    'img': f'screenshots/screenshot_{index+1}.png',
                    'code': status_code
                })
    finally:
        driver.quit()
    return screenshot_data

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files['file']
        column_name = request.form['column']
        if file:
            file_path = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(file_path)
            screenshot_data = open_urls_in_tabs(file_path, column_name)
            # Store screenshot_data in session or temp file, but for simplicity, pass as query string
            import json
            return redirect(url_for('dashboard', data=json.dumps(screenshot_data)))
    return render_template('index.html')
@app.route('/dashboard')
def dashboard():
    import json
    data = request.args.get('data', '')
    screenshot_data = json.loads(data) if data else []
    return render_template('dashboard.html', screenshot_data=screenshot_data)

if __name__ == '__main__':
    app.run(debug=True)